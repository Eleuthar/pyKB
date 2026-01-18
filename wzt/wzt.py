import io
from xlsxwriter import Workbook
from openpyxl import load_workbook
from datetime import date
from os import listdir
from dataclasses import dataclass
import json
from re import search
from copy import copy
from sys import argv, exit

                    
# customization
env = json.load(open('env.json'))
# first column dedicated for game count
COLUMN_OFFSET = env['column_offset'] + 66
BEGIN_ROUND_ROW = 4

@dataclass
class Member:
    nm: str
    bet_char: str
    done_char: str
    point_char: str
    bet: list
    done: list
    point: list
    report: list
    total: int


# keep user prompt on same line to prevent terminal clutter
def rewind_prompt(mzg, condition=None):
    opt = ''
    # prompt option must be digit in a certain range
    print()
    while True:
        while not opt.isdigit():
            print("\033[A\033[A")
            opt = input(f'{mzg}:  \b')
        opt = int(opt)
        if condition is not None:
            if not eval(condition):
                opt = ''
                continue
            else:
                return opt
        return opt


def group_count():
    gamer_num = ''
    while not gamer_num.isdigit():
        gamer_num = input('Numar jucatori: ')
    return int(gamer_num)


def join_players(gamer_num):
    group = []
    uzr_char = COLUMN_OFFSET
    for q in range(gamer_num):
        # bet \ done \ point
        bet_char = chr(uzr_char)
        done_char = chr(uzr_char + 1)
        point_char = chr(uzr_char + 2)
        uzr_char += 3
        who = input('Nume jucator: ')
        # nm,bet_char,done_char,point_char,winz,failz,total
        group.append(
            Member(
                nm=who,
                bet_char = bet_char,
                done_char = done_char,
                point_char = point_char,
                bet = [], done = [], point = [],
                report = [], total=0
            )
        )
    return group


# the amount of cards in one game depends on the number of players
def hand_num(gamer_num):
    return [1 for z in range(gamer_num)] + \
    [x for x in range(2,8)] + \
        [8 for z in range(gamer_num)] + \
            [x for x in range(7,1,-1)] + \
                [1 for z in range(gamer_num)]


def format_data(pen, formatting, pending_colorize, group, roundz, gamer_num):

    CHAR_DIFF = 65
    pen.set_row(0, 27)
    
    # game number column
    game_col = chr(COLUMN_OFFSET-1)
    pen.write(3, ord(game_col)-CHAR_DIFF, 'Nr', formatting['done'])
    
    # narrow game counter column
    pen.set_column(f"{game_col}:{game_col}", 5)

    # all rows except for name
    for row in range(1,4):
        pen.set_row(row, 25)
        group = [
            Member(nm='P1', bet_char='D', done_char='E', point_char='F',
                bet=[1,1,2,3,4,5,6,7,8,8,7,6,5,4,3,2,1,1],
                done = [0,0,2,0,4,5,6,7,8,8,7,6,5,4,3,0,0,0],
                point = [-1,-1,7,-3,9,10,11,12,13,13,12,11,10,9,8,-2,-1,-1],
                report=[], total=136
            ),
            Member(nm='P2', bet_char='D', done_char='E', point_char='F',
                bet=[1,1,2,3,4,5,6,7,8,8,7,6,5,4,3,2,1,1],
                done = [0,0,2,0,4,5,6,7,8,8,7,6,5,4,3,0,0,0],
                point = [-1,-1,7,-3,9,10,11,12,13,13,12,11,10,9,8,-2,-1,-1],
                report=[], total=136
            ),
        ]
    print(group)

    for uzr in group:
        try:
            # name
            pen.merge_range(f"{uzr.bet_char}1:{uzr.point_char}1", uzr.nm, formatting['header'])
            # total
            pen.merge_range(f"{uzr.bet_char}2:{uzr.point_char}2", uzr.total, formatting['total'])
            # stats
            for col in (uzr.bet_char, uzr.done_char, uzr.point_char):
                pen.set_column(f"{col}:{col}", 10)
            pen.write(2, ord(uzr.bet_char)-CHAR_DIFF, 'Pariat', formatting['stat'])
            pen.write(2, ord(uzr.done_char)-CHAR_DIFF, 'Facut', formatting['stat'])
            pen.write(2, ord(uzr.point_char)-CHAR_DIFF, 'Puncte', formatting['stat'])

            # game iteration from row 4
            for j in range(len(roundz)):
                row = j+3
                point = f'{uzr.point_char}{row}'
                color = pending_colorize.get(point, 'point')
                pen.set_row(row, 25)
                pen.write(row, ord(game_col)-CHAR_DIFF, f'#{j+1}', formatting['done'])
                pen.write(row, ord(uzr.bet_char)-CHAR_DIFF, uzr.bet[j], formatting['bet'])
                pen.write(row, ord(uzr.done_char)-CHAR_DIFF, uzr.done[j], formatting['done'])
                pen.write(row, ord(uzr.point_char)-CHAR_DIFF, uzr.point[j], formatting[color])
        except Exception as x:
            print(x)
            

    # table bottom border
    # pen.merge_range(
    #     f'{group[0].bet_char}{row+1}:{group[-1].point_char}{row+1}', 
    #     '', wb.add_format({'top': 2}))


def prompt_bet(who, ndx, final_bidder, bid, hand):
    turn = f'{who} -->'
    allowed = None
    # no constraint for final bidder if bid > hand
    if ndx == final_bidder and bid <= hand:
        diff = hand - bid
        allowed = [q for q in range(0, hand+1)]
        allowed.remove(diff)
    else:
        allowed = [q for q in range(0, hand+1)]
    mzg = f"{turn} {allowed}"
    condition = f'opt in {allowed}'
    bet = int(rewind_prompt(mzg, condition=condition))
    return bet


def next_fight():
    """ Return workbook filename """
    ROUND = 1
    next_fight = [
        int(fname.split()[-1].split('.')[0]) 
        for fname in listdir() if 'ROUND' in fname
    ]
    if len(next_fight) > 0:
        ROUND = max(next_fight)+1
    fname = f"ROUND {ROUND}.xlsx"
    return fname, ROUND


# mark Cell for color formatting during datapen dump
def colorize(pending_colorize, color, point_char, row, gamer_num):
    for j in range(row-4, row+1):
        pending_colorize[f'{point_char}{j}'] = f'{color}_point'
    return pending_colorize


def parse_point(gamer_num, game, hand, bidder, colorize, pending_colorize):
    BONUS = 5
    BEGIN_BONUS_ROUND = gamer_num + BONUS
    row = game + BEGIN_ROUND_ROW
    done = bidder.done[game]
    bet = bidder.bet[game]
    # win
    if done == bet:
        point = 5 + bet
        bidder.point.append(point)
        bidder.total += bidder.point[game]
        if hand > 1:
            bidder.report.append(1)
            # positive bonus & reset streak
            if game >= BEGIN_BONUS_ROUND:
                winz = bidder.report[-1:-6:-1].count(1)
                if winz == BONUS:
                    bidder.report = []
                    bidder.total += (BONUS * gamer_num)
                    pending_colorize = colorize(
                        pending_colorize, 'green', bidder.point_char, row-1, gamer_num)
    # lose
    else:
        # make negative to positive to allow subtraction from total
        point = int(str(done - bet).strip('-'))
        bidder.point.append(f'-{point}')
        bidder.total -= point
        if hand > 1:
            bidder.report.append(0)
            if game >= BEGIN_BONUS_ROUND:
                print(bidder.report)
                failz = bidder.report[-1:-6:-1].count(0)
                if failz == BONUS:
                    bidder.report = []
                    # negative bonus & reset streak
                    bidder.total -= (BONUS * gamer_num)
                    pending_colorize = colorize(
                        pending_colorize, 'red', bidder.point_char, row-1, gamer_num)
    return pending_colorize


if __name__ == '__main__':
    # in memory writer object
    output = io.BytesIO()
    gamer_num = group_count()
    group = join_players(gamer_num)
    # number of deck dealing per game
    roundz = hand_num(gamer_num)
    # find previous score workbooks in game directory
    fname, ROUND = next_fight()
    wb = Workbook(fname)
    pen = wb.add_worksheet(f'ROUND {ROUND}')
    # formatting
    pending_colorize = {}
    formatting = {}
    for bg in ['header','total','stat','bet','done','point','green_point','red_point']:
        formatting[bg] = wb.add_format(env['formatting'][bg])
    _next = -1

    while True:
        print(f'\nSpor la joaca!\n')
        for game in range(len(roundz)):
            row = game+4
            hand = roundz[game]
            print(f"\n\n#{game+1} Runda de {hand}\n{'='*len('runda de xxxx')}")

            # bidding
            print(f'\nPariaza\n{"`"*len("nPariaza")}')
            bid = 0
            _next += 1
            if _next == gamer_num:
                _next = 0
            order = [x for x in range(_next, gamer_num)] + [x for x in range(0, _next)]
            final_bidder = order[-1]
            for ndx in order:
                who = group[ndx]
                bet = prompt_bet(who.nm, ndx, final_bidder, bid, hand)
                who.bet.append(bet)
                bid += bet
            
            # done
            print(f'\nMaini facute\n{"`"*len("Maini facute")}')
            done_total = 0
            for ndx in order:
                bidder = group[ndx]
                condition = f"{done_total} + opt <= {hand} and opt <= {hand}"
                done = rewind_prompt(bidder.nm, condition)
                done_total += done
                bidder.done.append(done)
                # winner
                bet = bidder.bet[-1]
                # update dict of cells to be colored
                pending_colorize = parse_point(
                    gamer_num, game, hand, bidder, colorize, pending_colorize)
        format_data(pen, formatting, pending_colorize, group, roundz, gamer_num)
        wb.close()
        
        _next = ''
        while not _next.isalpha():
            _next = input("\n\nJoc nou? Y \\ N: ")
            _next = _next.upper()
            if _next.upper() not in ['Y', 'N']:
                continue
            elif _next == 'N':
                exit()
            else:
                _next = -1
                ROUND += 1
                output = io.BytesIO()
                fname = f'ROUND {ROUND}.xlsx'
                wb = Workbook(output, {'in_memory': True})
                pen = wb.add_worksheet(f'ROUND {ROUND}')
                for uzr in group:
                    for prop in ['total','winz','failz']:
                        setattr(uzr, prop, 0)
                    for prop in ['bet','done','point']:
                        setattr(uzr, prop, [])
                break
