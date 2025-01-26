from openpyxl import load_workbook
from copy import copy


xxpath='REG.xlsx'
workbook = load_workbook(xxpath, data_only=True)


# copy cell formatting from in-memory workbook
def export_dataframe(output, fname, ROUND, max_col, max_row, begin_char):
    output.seek(0)
    memwb = load_workbook(output)
    mainframe = memwb.active
    tgt_wb = load_workbook(fname)
    tgt_frame = tgt_wb.create_sheet(f"ROUND {ROUND}")
    # column width
    for col in range(begin_char, max_col):
        char = chr(col)
        tgt_frame.column_dimensions[char].width = mainframe.column_dimensions[char].width
    for row in range(1, max_row):
        # row height
        tgt_frame.row_dimensions[row].height = mainframe.row_dimensions[row].height
        for col in range(begin_char, max_col):
            coord = f'{chr(col)}{row}'
            orig = mainframe[coord]
            tgt = tgt_frame[coord]
            tgt.value = orig.value
            tgt.font = copy(orig.font)
            tgt.fill = copy(orig.fill)
            tgt.border = copy(orig.border)
            tgt.alignment = copy(orig.alignment)
            tgt_wb.save(fname)
            output.close()
            tgt.data_type='n'
            tgt.number_format='#,##0.00'
            tgt.value = 0
    # MERGED last due to readonly error
    for merged_range in mainframe.merged_cells.ranges:
        tgt_frame.merge_cells(str(merged_range))
    tgt_wb.save(fname)
    output.close()
    

# generate frame cell coordinate
def gen_coord(frame):
    for x in range(18, 20):
        for q in range(ord('A'), ord('K')):
            q = chr(q)
            tgt = frame[f'{q}{x}']
        
