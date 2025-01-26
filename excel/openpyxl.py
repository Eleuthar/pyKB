from openpyxl import load_workbook
from copy import copy


xxpath='REG.xlsx'
workbook = load_workbook(xxpath, data_only=True)


# copy cell formatting from in-memory workbook

frame.column_dimensions['A'].width = 27
frame.insert_rows(9,1)
frame.row_dimensions[x].height = 22.5
def export_dataframe(output, fname, max_col, max_row, begin_char):
    output.seek(0)
    memwb = load_workbook(output)
    mainframe = memwb.active
    tgt_wb = load_workbook(fname)
    tgt_frame = tgt_wb.create_sheet(f"ROUND {ROUND}")
    # from pdb import set_trace
    # set_trace()
    for row in range(max_row):
        for col in range(begin_char, max_col):
            coord = f'{chr(col)}{chr(row)}'
            orig = mainframe[coord]
            tgt = tgt_frame[coord]
            tgt.font = copy(orig.font)
            tgt.fill = copy(orig.fill)
            tgt.border = copy(orig.border)
            tgt.alignment = copy(orig.alignment)
            tgt.font = copy(orig.font)
            tgt.border = copy(orig.border)
            tgt.alignment = copy(orig.alignment)
            tgt.data_type='n'
            tgt.number_format='#,##0.00'
            tgt.value = 0
    tgt.save(fname)
    output.close()
    

# generate frame cell coordinate
for x in range(18, 20):
    for q in range(ord('A'), ord('K')):
        q = chr(q)
        tgt = frame[f'{q}{x}']
        
