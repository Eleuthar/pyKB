import openpyxl
from openpyxl.styles import Alignment

# Helper function to merge and align cells
def merge_and_write(sheet, start_row, end_row, start_col, end_col, value):
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = sheet.cell(row=start_row, column=start_col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")

z = workbook.create_sheet(title=f"Sheet{len(workbook.sheetnames)+1}")
merge_and_write(z, 1, 1, 1, 4, 'PAROHIA DOMUS - VOLUNTARI')
merge_and_write(z, 2, 2, 1, 4, 'REGISTRU LUMANARI \ COLPORTAJ PANGAR')

# Add headers
merge_and_write(z, 5, 8, 1, 1, "TIP PRODUS")
merge_and_write(z, 5, 5, 2, 6, "INTRARI")
merge_and_write(z, 5, 5, 7, 8, "IESIRI")
merge_and_write(z, 5, 5, 9, 10, "STOCURI")

# INTRARI subheader
merge_and_write(z, 6, 7, 2, 3, "Cantitate")
merge_and_write(z, 7, 8, 2, 2, "Adaugat")
merge_and_write(z, 7, 8, 3, 3, "Anterior")

merge_and_write(z, 6, 8, 4, 4, "U.M.")

merge_and_write(z, 6, 6, 5, 5, "Pret unitar")
merge_and_write(z, 7, 7, 5, 5, "")
merge_and_write(z, 8, 8, 5, 5, "~ LEI ~")

merge_and_write(z, 6, 6, 6, 6, "Valoare totala")
merge_and_write(z, 7, 7, 6, 6, "(col. 1 x (col. 2 + 3))")
merge_and_write(z, 8, 8, 6, 6, "~ LEI ~")

# IESIRI subheader
merge_and_write(z, 6, 8, 7, 7, "Cantitate")
merge_and_write(z, 6, 6, 8, 8, "Valoare totala")
merge_and_write(z, 7, 7, 8, 8, "(col. 3 x col. 5)")
merge_and_write(z, 8, 8, 8, 8, "~ LEI ~")

# STOCURI subheader
merge_and_write(z, 6, 6, 9, 9, "Cantitate")
merge_and_write(z, 7, 7, 9, 9, "(col. 1 - col. 5)")
merge_and_write(z, 8, 8, 9, 9, "")
merge_and_write(z, 6, 6, 10, 10, "Valoare")
merge_and_write(z, 7, 7, 10, 10, "(col. 3 x col. 7)")
merge_and_write(z, 8, 8, 10, 10, "~ LEI ~")

# index
ndx=0
for x in range(ord('A'), ord('K')):
    z[f'{chr(x)}9'] = ndx
    ndx+=1

# item
z['A10'] = 'Lumanari 100B'
z['A11'] = 'Lumanari C20'
z['A12'] = 'Candele tip 0'
z['A13'] = 'Candele tip 1'
z['A14'] = 'Candele tip 2'
z['A15'] = 'Candele tip 3'
z['A16'] = 'Candele tip 4'
z['A17'] = 'COLPORTAJ'
z['A18'] = 'TOTAL'

# C \\ Importare Stoc din sheet anterior I10:16  
z[f'C{x}'].value = prev[f'I{x}'].value
