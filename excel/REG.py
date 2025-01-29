from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment



"""
Generator de formule pt registre saptamanale + raport general
"""


# start row of products under weekly registry
PROD_OFFSET = 9
BEGIN_ROW = 10
END_ROW = 16
TOTAL_ROW = END_ROW + 1

prod = {
    1: {'prod':'Lumanari 100B', '$': 0, 'chr': 'D'},
    2: {'prod':'Lumanari C20', '$': 0, 'chr': 'E'},
    3: {'prod':'Candele tip 0', '$': 0, 'chr': 'F'},
    4: {'prod':'Candele tip 1', '$': 0, 'chr': 'G'},
    5: {'prod':'Candele tip 2', '$': 0, 'chr': 'H'},
    6: {'prod':'Candele tip 3', '$': 0, 'chr': 'I'},
    7: {'prod':'Candele tip 4', '$': 0, 'chr': 'J'}
}


xxpath = 'REG.xlsx'
workbook = load_workbook(xxpath)
max_ndx = len(workbook.worksheets)
report = workbook.worksheets[0]
year = 2024
# intrare cantitate\total, iesire cantitate\total, stoc cantitate\total


# sheet range for general report formulas
month_mapping = {
    'IAN': { 'row': 9, 'range': '', 'wkz': [] }, 
    'FEB': { 'row': 15, 'range': '', 'wkz': [] }, 
    'MAR': { 'row': 21, 'range': '', 'wkz': [] }, 
    'APR': { 'row': 27, 'range': '', 'wkz': [] },
    'MAI': { 'row': 33, 'range': '', 'wkz': [] }, 
    'IUN': { 'row': 39, 'range': '', 'wkz': [] }, 
    'IUL': { 'row': 45, 'range': '', 'wkz': [] }, 
    'AUG': { 'row': 51, 'range': '', 'wkz': [] }, 
    'SEPT': { 'row': 57, 'range': '', 'wkz': [] }, 
    'OCT': { 'row': 63, 'range': '', 'wkz': [] }, 
    'NOV': { 'row': 69, 'range': '', 'wkz': [] },
    'DEC': { 'row': 75, 'range': '', 'wkz': [] }
}


# Helper function to merge and align cells
def merge_and_write(sheet, start_row, end_row, start_col, end_col, value):
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = sheet.cell(row=start_row, column=start_col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def gen_week_sheet(workbook,fname):
    z = workbook.create_sheet(title=fname)
    merge_and_write(z, 1, 1, 1, 4, 'PAROHIA DOMUS - VOLUNTARI')
    merge_and_write(z, 2, 2, 1, 4, 'REGISTRU LUMANARI \ COLPORTAJ PANGAR')

    # Add headers
    merge_and_write(z, 5, 8, 1, 1, "TIP PRODUS")
    merge_and_write(z, 5, 5, 2, 6, "INTRARI")
    merge_and_write(z, 5, 5, 7, 8, "IESIRI")
    merge_and_write(z, 5, 5, 9, 10, "STOCURI")

    # INTRARI subheader
    merge_and_write(z, 6, 7, 2, 3, "Cantitate")
    merge_and_write(z, 7, 8, 2, 2, "Adaugat")
    merge_and_write(z, 7, 8, 3, 3, "Anterior")

    merge_and_write(z, 6, 8, 4, 4, )

    merge_and_write(z, 6, 6, 5, 5, "Pret unitar")
    merge_and_write(z, 7, 7, 5, 5, "")
    merge_and_write(z, 8, 8, 5, 5, "~ LEI ~")

    merge_and_write(z, 6, 6, 6, 6, "Valoare totala")
    merge_and_write(z, 7, 7, 6, 6, "(col. 1 x (col. 2 + 3))")
    merge_and_write(z, 8, 8, 6, 6, "~ LEI ~")

    # IESIRI subheader
    merge_and_write(z, 6, 8, 7, 7, "Cantitate")
    merge_and_write(z, 6, 6, 8, 8, "Valoare totala")
    merge_and_write(z, 7, 7, 8, 8, "(col. 3 x col. 5)")
    merge_and_write(z, 8, 8, 8, 8, "~ LEI ~")

    # STOCURI subheader
    merge_and_write(z, 6, 6, 9, 9, "Cantitate")
    merge_and_write(z, 7, 7, 9, 9, "(col. 1 - col. 5)")
    merge_and_write(z, 8, 8, 9, 9, "")
    merge_and_write(z, 6, 6, 10, 10, "Valoare")
    merge_and_write(z, 7, 7, 10, 10, "(col. 3 x col. 7)")
    merge_and_write(z, 8, 8, 10, 10, "~ LEI ~")
    # product index
    ndx=0
    for x in range(ord('A'), ord('K')):
        z[f'{chr(x)}9'] = ndx
        ndx+=1
    # item
    z['A10'] = 'Lumanari 100B'
    z['A11'] = 'Lumanari C20'
    z['A12'] = 'Candele tip 0'
    z['A13'] = 'Candele tip 1'
    z['A14'] = 'Candele tip 2'
    z['A15'] = 'Candele tip 3'
    z['A16'] = 'Candele tip 4'
    z['A17'] = 'COLPORTAJ'
    z['A18'] = 'TOTAL'


def find_negative(wb_name, col_range, begin=BEGIN_ROW, end=TOTAL_ROW):
    '''col_range = ['B','C','F','G','H','I','J']'''
    workbook = load_workbook(wb_name, data_only=True)
    for fm in workbook.worksheets[1:]:
        for prod in col_range:
            for row in range(begin, end):
                try:
                    tgt = fm[f'{prod}{row}']
                    if tgt.value < 0:
                        print(fm.title, prod, row, tgt.value)
                except:
                    try:
                        if tgt.value < 0:
                            print(fm.title, prod, row, tgt.value)
                    except:
                        print(tgt.value)


# make sheet title match the date_cell
def rename_title(workbook, date_cell):
    for frame in workbook.worksheets[1:]:
        dt = frame[date_cell]
        head = dt.value.split(":")
        part = head[1].lstrip(" 0").upper()
        head[1] = part
        dt.value = ': '.join(head)
        frame.title = ' '.join(part.split()[:-1])


# helper function to ensure the formula referencing the previous sheet matches the title
# charz is a list of columns that hold the reference
def match_form_with_prev_title(workbook, prev_char, tgt_char, begin_row, end_row, fix=False):
    # exclude first sheet as the report & 2nd sheet as formula origin
    for x in range(2, len(workbook.sheetnames)):
        prev = workbook.sheetnames[x-1]
        frame = workbook.worksheets[x]
        for row in range(begin_row, end_row):
            tgt = frame[f'{tgt_char}{row}']
            expected = f"='{prev}'!{prev_char}"
            if tgt.value != expected:
                print(f'FOUND {tgt.value}, {tgt_char}{row}, {prev} < Apply fix? ')
                if fix:
                    tgt.value = expected
                    
# prevent NoneType error
def init_report(report, begin_chr='D', end_chr='J', begin_row=9, end_row=84):
    for col in range(ord(begin_chr), ord(end_chr)+1):
        for row in range(begin_row, end_row+1):
            report[f'{chr(col)}{row}'].value = ''


# for building formula by `generate_monthly_report_formula`
def gather_month_wkz(workbook, month_mapping):
    wb_ndx = 1
    for month in month_mapping:
        for wk in workbook.sheetnames[wb_ndx:]:
            if month in wk:
                month_mapping[month]['wkz'].append(wk)
                wb_ndx += 1
            else:
                # update range value
                begin = month_mapping[month]['wkz'][0]
                end = month_mapping[month]['wkz'][-1]
                month_mapping[month]['range'] = f'{begin}:{end}'
                break
    # handle December, having no next month to trigger else condition
    begin = month_mapping[month]['wkz'][0]
    end = month_mapping[month]['wkz'][-1]
    month_mapping[month]['range'] = f'{begin}:{end}'


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ GENERAL REPORT BUILDER ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

def generate_monthly_report_formula(month_mapping, month, prod_row):

    # =('1 IAN'!E10 * '1 IAN'!B10) + ('8 IAN'!E10 * '8 IAN'!B10) + ...
    amount_in = '='
    for wk in month_mapping[month]['wkz']:
        amount_in += f"('{wk}'!E{prod_row} * '{wk}'!B{prod_row}) + "
    amount_in = amount_in.rstrip(' + ')
    wk_range = month_mapping[month]['range']
    end_wk = month_mapping[month]['wkz'][-1]

    # =SUM('1 IAN:29 IAN'!B10)
    quant_in = f"=SUM('{wk_range}'!{'B'}{prod_row})"
    quant_out = f"=SUM('{wk_range}'!{'G'}{prod_row})"
    amount_out = f"=SUM('{wk_range}'!{'H'}{prod_row})"

    # stock is using the last sheet from month_mapping['range']
    quant_stock = f"='{end_wk}'!{'I'}{prod_row}"
    amount_stock = f"='{end_wk}'!{'J'}{prod_row}"
    return [quant_in, amount_in, quant_out, amount_out, quant_stock, amount_stock]


def report_form(PROD_OFFSET, prod, report, month_mapping):
    # gather each month report row relevant to general totals
    quant_in_month_row = []
    amount_in_month_row = []
    quant_out_month_row = []
    amount_out_month_row = []
    for month in month_mapping:
        report_row = month_mapping[month]['row']
        quant_in_month_row.append(report_row)
        amount_in_month_row.append(report_row+1)
        quant_out_month_row.append(report_row+2)
        amount_out_month_row.append(report_row+3)

    for prod_ndx in prod:
        prod_row = prod_ndx + PROD_OFFSET        
        for month in month_mapping:
            report_row = month_mapping[month]['row']
            # store the row number for SUM aggregation in general totals
            # monthly formulas
            prod_chr = prod[prod_ndx]['chr']
            quant_in = report[f'{prod_chr}{report_row}']
            amount_in = report[f'{prod_chr}{report_row+1}']
            quant_out = report[f'{prod_chr}{report_row+2}']
            amount_out = report[f'{prod_chr}{report_row+3}']
            quant_stock = report[f'{prod_chr}{report_row+4}']
            amount_stock = report[f'{prod_chr}{report_row+5}']
            [
                quant_in.value,
                amount_in.value,
                quant_out.value,
                amount_out.value,
                quant_stock.value,
                amount_stock.value
            ] = generate_monthly_report_formula(month_mapping, month, prod_row)

    # GENERAL TOTALS   
    # generate sum formula after month rows have been gathered
    for prod_ndx in prod:
        prod_chr = prod[prod_ndx]['chr']
        
        range_quant_in = [f'{prod_chr}{x}' for x in quant_in_month_row]
        report[f'{prod_chr}81'].value = "=SUM(" + ', '.join(range_quant_in) + ")"

        range_amount_in = [f'{prod_chr}{x}' for x in amount_in_month_row]
        report[f'{prod_chr}82'].value = "=SUM(" + ', '.join(range_amount_in) + ")"

        range_quant_out = [f'{prod_chr}{x}' for x in quant_out_month_row]
        report[f'{prod_chr}83'].value = "=SUM(" + ', '.join(range_quant_out) + ")"

        range_amount_out = [f'{prod_chr}{x}' for x in amount_out_month_row]
        report[f'{prod_chr}84'].value = "=SUM(" + ', '.join(range_amount_out) + ")"


if __name__ == '__main__':
    init_report(report)
    gather_month_wkz(workbook, month_mapping)
    # helper `generate_monthly_report_formula(month_mapping, month, prod_row)`
    report_form(PROD_OFFSET, prod, report, month_mapping)
    workbook.save(xxpath)
    workbook.close()