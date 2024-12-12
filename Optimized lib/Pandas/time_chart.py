import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


# Create the dataset with the Incident column
data_with_incidents = {
    "Time": ["9:00 AM", "9:10 AM", "9:20 AM", "9:30 AM", "9:40 AM", "9:50 AM", "10:00 AM"],
    "Value": [10, 20, 30, 40, 50, 60, 70],
    "Incident": ["Incident A", "", "Incident B", "Incident A", "", "Incident C", ""]
}
df_incidents = pd.DataFrame(data_with_incidents)

# Create an Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Time Data"

# Write the dataframe to the sheet
for r_idx, row in enumerate(dataframe_to_rows(df_incidents, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Add an Excel table
table = Table(displayName="TimeData", ref=f"A1:C{len(df_incidents) + 1}")
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True)

table.tableStyleInfo = style
ws.add_table(table)

# Add a drop-down validation for incidents
validation_col = ws.cell(row=1, column=5)
validation_col.value = "Selected Incident"

# Add unique incidents for dropdown list
incident_list = [incident for incident in df_incidents["Incident"].unique() if incident]
for i, incident in enumerate(incident_list, start=1):
    ws.cell(row=i + 1, column=5, value=incident)

# Save the Excel file
excel_file_path = "/mnt/data/Time_Based_Chart_with_Incidents.xlsx"
wb.save(excel_file_path)